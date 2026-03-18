"""
Generate multi-period strategy vs benchmark comparison Excel workbook.
Periods: Full (1990), 2000, 2010, 2020.
"""
from __future__ import annotations

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path

from data_fetcher import fetch_all
from market_cap_estimator import estimate_market_caps, rank_by_market_cap
from backtest_engine import run_backtest, make_top1_fn, make_topn_fn, make_momentum_fn
from metrics import compute_metrics


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PERIODS = {
    "Full (1990–2026)": "1990-01-31",
    "2000–2026": "2000-01-31",
    "2010–2026": "2010-01-31",
    "2020–2026": "2020-01-31",
}

# STRATEGIES dict is created inside main() to avoid module-level side effects.

METRIC_KEYS = [
    "cagr_twr",
    "total_return_twr",
    "annualized_volatility",
    "sharpe_ratio",
    "sortino_ratio",
    "max_drawdown",
    "max_drawdown_duration_days",
    "calmar_ratio",
    "omega_ratio",
]

METRIC_LABELS = {
    "cagr_twr": "CAGR (TWR)",
    "total_return_twr": "Total Return",
    "annualized_volatility": "Volatility (Ann.)",
    "sharpe_ratio": "Sharpe Ratio",
    "sortino_ratio": "Sortino Ratio",
    "max_drawdown": "Max Drawdown",
    "max_drawdown_duration_days": "Max DD Duration (days)",
    "calmar_ratio": "Calmar Ratio",
    "omega_ratio": "Omega Ratio",
}

# Metrics where higher is worse (for conditional formatting)
LOWER_IS_BETTER = {"annualized_volatility", "max_drawdown", "max_drawdown_duration_days"}
# Metrics displayed as percentages
PCT_METRICS = {"cagr_twr", "total_return_twr", "annualized_volatility", "max_drawdown"}
# Metrics displayed as integers
INT_METRICS = {"max_drawdown_duration_days"}

# Styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
STRATEGY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BENCH_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WORST_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
DATA_FONT = Font(name="Calibri", size=11)
MONEY_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color="1F4E79"))


def clean_twr(twr: pd.Series) -> pd.Series:
    if len(twr) > 1 and twr.iloc[0] == 0:
        return twr.iloc[1:]
    return twr


def run_period(data, mcaps, rankings, rf, benchmarks, start_str):
    """Run all strategies and benchmarks for a given start date."""
    start = pd.Timestamp(start_str)
    results = {}

    # Strategies
    for name, fn in STRATEGIES.items():
        res = run_backtest(
            strategy_fn=fn, prices=data["prices"], rankings=rankings,
            market_caps=mcaps, risk_free=rf,
            initial_capital=10000, monthly_contribution=1000,
            cost_schedule=None, start_date=start_str, end_date=None,
            strategy_name=name,
        )
        twr = clean_twr(res.twr_returns)
        rfr = rf.reindex(twr.index).fillna(0) if len(rf) > 0 else pd.Series(0, index=twr.index)
        m = compute_metrics(twr, res.cash_flows, res.equity_curve, pd.Series(dtype=float), rfr, res.holdings_history)

        months = len(res.equity_curve)
        total_contrib = 10000 + 1000 * max(0, months - 1)
        results[name] = {
            "metrics": m,
            "final_value": res.equity_curve.iloc[-1],
            "total_contributed": total_contrib,
            "months": months,
            "start": res.equity_curve.index[0],
            "end": res.equity_curve.index[-1],
            "is_strategy": True,
            "backtest_result": res,
        }

    # Benchmarks
    ref = next(iter(results.values()))
    ref_start = ref["start"]
    ref_end = ref["end"]

    for bname in benchmarks["benchmark"].unique():
        bp = benchmarks[benchmarks["benchmark"] == bname][["date", "adj_close"]].copy()
        bp["date"] = pd.to_datetime(bp["date"])
        bp = bp.set_index("date").sort_index()["adj_close"]
        bret = bp.pct_change().dropna()
        bret = bret[(bret.index >= ref_start) & (bret.index <= ref_end)]

        if len(bret) < 6:
            continue

        # Build equity with contributions
        prev = 10000.0
        eq_vals, eq_dates = [], []
        for j, (dt, r) in enumerate(bret.items()):
            contrib = 1000.0 if j > 0 else 0.0
            new_val = prev * (1 + r) + contrib
            eq_vals.append(new_val)
            eq_dates.append(dt)
            prev = new_val
        bench_eq = pd.Series(eq_vals, index=eq_dates)

        # Cash flows
        cf_rows = [{"date": bret.index[0], "amount": -10000}]
        for d in bret.index[1:]:
            cf_rows.append({"date": d, "amount": -1000})
        cf_rows.append({"date": bret.index[-1], "amount": bench_eq.iloc[-1]})
        cf = pd.DataFrame(cf_rows)

        rfr = rf.reindex(bret.index).fillna(0) if len(rf) > 0 else pd.Series(0, index=bret.index)
        bm = compute_metrics(bret, cf, bench_eq, pd.Series(dtype=float), rfr)

        months = len(bret)
        total_contrib = 10000 + 1000 * max(0, months - 1)
        results[bname] = {
            "metrics": bm,
            "final_value": bench_eq.iloc[-1],
            "total_contributed": total_contrib,
            "months": months,
            "start": bret.index[0],
            "end": bret.index[-1],
            "is_strategy": False,
        }

    return results


def write_period_sheet(wb: Workbook, sheet_name: str, period_label: str, results: dict):
    """Write one period's comparison to a worksheet."""
    ws = wb.create_sheet(title=sheet_name)
    series_names = list(results.keys())
    n_series = len(series_names)

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_series)
    ws.cell(1, 1, f"Strategy vs Benchmark Comparison — {period_label}").font = TITLE_FONT

    ref = next(iter(results.values()))
    subtitle = f"{ref['start'].strftime('%b %Y')} – {ref['end'].strftime('%b %Y')}  |  $10K initial + $1K/mo contributions  |  Strategy returns net of transaction costs"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + n_series)
    ws.cell(2, 1, subtitle).font = SUBTITLE_FONT

    # --- Metrics Table ---
    table_start = 4
    # Header row
    ws.cell(table_start, 1, "Metric").font = HEADER_FONT
    ws.cell(table_start, 1).fill = HEADER_FILL
    ws.cell(table_start, 1).alignment = Alignment(horizontal="left")
    for j, name in enumerate(series_names):
        cell = ws.cell(table_start, 2 + j, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, k in enumerate(METRIC_KEYS):
        row = table_start + 1 + i
        label_cell = ws.cell(row, 1, METRIC_LABELS.get(k, k))
        label_cell.font = Font(name="Calibri", size=11, bold=True)
        label_cell.border = THIN_BORDER

        values = []
        for j, name in enumerate(series_names):
            v = results[name]["metrics"].get(k, float("nan"))
            values.append(v)
            cell = ws.cell(row, 2 + j)

            if v is None or (isinstance(v, float) and np.isnan(v)):
                cell.value = "—"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = v
                cell.alignment = Alignment(horizontal="center")
                if k in PCT_METRICS:
                    cell.number_format = "0.00%"
                elif k in INT_METRICS:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "0.0000"

            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            # Strategy vs benchmark background
            if results[name]["is_strategy"]:
                cell.fill = STRATEGY_FILL
            else:
                cell.fill = BENCH_FILL

        # Highlight best/worst
        numeric_vals = [(j, v) for j, v in enumerate(values)
                        if v is not None and not (isinstance(v, float) and np.isnan(v))]
        if len(numeric_vals) >= 2:
            if k in LOWER_IS_BETTER:
                best_j = min(numeric_vals, key=lambda x: x[1])[0]
                worst_j = max(numeric_vals, key=lambda x: x[1])[0]
            else:
                best_j = max(numeric_vals, key=lambda x: x[1])[0]
                worst_j = min(numeric_vals, key=lambda x: x[1])[0]

            ws.cell(row, 2 + best_j).fill = BEST_FILL
            ws.cell(row, 2 + worst_j).fill = WORST_FILL

    # --- Final Values Section ---
    val_start = table_start + len(METRIC_KEYS) + 3
    ws.merge_cells(start_row=val_start, start_column=1, end_row=val_start, end_column=1 + n_series)
    ws.cell(val_start, 1, "Portfolio Summary").font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    summary_headers = ["", "Final Value", "Total Contributed", "Profit", "Multiple"]
    # Transpose: series as rows
    for j, h in enumerate(summary_headers):
        cell = ws.cell(val_start + 1, 1 + j, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, name in enumerate(series_names):
        row = val_start + 2 + i
        r = results[name]
        profit = r["final_value"] - r["total_contributed"]
        mult = r["final_value"] / r["total_contributed"]

        ws.cell(row, 1, name).font = Font(name="Calibri", size=11, bold=True)
        if r["is_strategy"]:
            ws.cell(row, 1).fill = STRATEGY_FILL
        else:
            ws.cell(row, 1).fill = BENCH_FILL

        for j, (val, fmt) in enumerate([
            (r["final_value"], '"$"#,##0'),
            (r["total_contributed"], '"$"#,##0'),
            (profit, '"$"#,##0'),
            (mult, '0.0"x"'),
        ]):
            cell = ws.cell(row, 2 + j, val)
            cell.number_format = fmt
            cell.font = MONEY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            if r["is_strategy"]:
                cell.fill = STRATEGY_FILL
            else:
                cell.fill = BENCH_FILL

    # Highlight best final value
    final_vals = [(i, results[name]["final_value"]) for i, name in enumerate(series_names)]
    best_i = max(final_vals, key=lambda x: x[1])[0]
    ws.cell(val_start + 2 + best_i, 2).fill = BEST_FILL

    # Highlight best multiple
    mult_vals = [(i, results[name]["final_value"] / results[name]["total_contributed"])
                 for i, name in enumerate(series_names)]
    best_m = max(mult_vals, key=lambda x: x[1])[0]
    ws.cell(val_start + 2 + best_m, 5).fill = BEST_FILL

    # --- Column widths ---
    ws.column_dimensions["A"].width = 26
    for j in range(n_series):
        col_letter = get_column_letter(2 + j)
        ws.column_dimensions[col_letter].width = 20

    # Freeze panes
    ws.freeze_panes = "B5"


def _derive_position_changes(holdings_history: pd.DataFrame, strategy_name: str) -> pd.DataFrame:
    """
    Derive ENTER/EXIT events by comparing consecutive months in holdings_history.

    Returns DataFrame with columns: date, strategy, ticker, action, weight.
    """
    if holdings_history.empty:
        return pd.DataFrame(columns=["date", "strategy", "ticker", "action", "weight"])

    records = []
    dates = sorted(holdings_history["date"].unique())
    prev_tickers = set()
    for date in dates:
        month_df = holdings_history[holdings_history["date"] == date]
        curr_tickers = set(month_df["ticker"].values)
        weight_map = dict(zip(month_df["ticker"], month_df["weight"]))

        entered = curr_tickers - prev_tickers
        exited = prev_tickers - curr_tickers

        for t in sorted(entered):
            records.append({
                "date": date, "strategy": strategy_name,
                "ticker": t, "action": "ENTER", "weight": weight_map.get(t, 0),
            })
        for t in sorted(exited):
            records.append({
                "date": date, "strategy": strategy_name,
                "ticker": t, "action": "EXIT", "weight": 0.0,
            })
        prev_tickers = curr_tickers

    return pd.DataFrame(records)


def write_holdings_sheet(wb: Workbook, sheet_name: str, period_label: str, results: dict):
    """Write a Holdings sheet for a given period showing current holdings, full timeline, and position changes."""
    ws = wb.create_sheet(title=sheet_name)
    row = 1

    # ===== Section A: Current Holdings (most recent date) =====
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, f"Current Holdings — {period_label}").font = TITLE_FONT
    row += 2

    strategy_results = {k: v for k, v in results.items() if v.get("is_strategy") and v.get("backtest_result")}

    for sname, sdata in strategy_results.items():
        hh = sdata["backtest_result"].holdings_history
        if hh.empty:
            continue
        latest_date = hh["date"].max()
        latest = hh[hh["date"] == latest_date].sort_values("weight", ascending=False)

        # Strategy header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, f"{sname} — {latest_date.strftime('%Y-%m')}").font = Font(
            name="Calibri", size=11, bold=True, color="1F4E79"
        )
        ws.cell(row, 1).fill = STRATEGY_FILL
        row += 1

        # Column headers
        for j, h in enumerate(["Ticker", "Shares", "Value ($)", "Weight (%)"]):
            cell = ws.cell(row, 1 + j, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows
        total_value = 0.0
        for _, hr in latest.iterrows():
            ws.cell(row, 1, hr["ticker"]).font = DATA_FONT
            ws.cell(row, 1).border = THIN_BORDER

            cell_shares = ws.cell(row, 2, hr["shares"])
            cell_shares.number_format = "#,##0.00"
            cell_shares.font = DATA_FONT
            cell_shares.alignment = Alignment(horizontal="center")
            cell_shares.border = THIN_BORDER

            cell_val = ws.cell(row, 3, hr["value"])
            cell_val.number_format = '"$"#,##0'
            cell_val.font = MONEY_FONT
            cell_val.alignment = Alignment(horizontal="center")
            cell_val.border = THIN_BORDER

            cell_wt = ws.cell(row, 4, hr["weight"])
            cell_wt.number_format = "0.00%"
            cell_wt.font = DATA_FONT
            cell_wt.alignment = Alignment(horizontal="center")
            cell_wt.border = THIN_BORDER

            total_value += hr["value"]
            row += 1

        # Total row
        ws.cell(row, 1, "TOTAL").font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row, 1).border = BOTTOM_BORDER
        cell_total = ws.cell(row, 3, total_value)
        cell_total.number_format = '"$"#,##0'
        cell_total.font = MONEY_FONT
        cell_total.alignment = Alignment(horizontal="center")
        cell_total.border = BOTTOM_BORDER
        cell_wt_total = ws.cell(row, 4, 1.0)
        cell_wt_total.number_format = "0.00%"
        cell_wt_total.font = Font(name="Calibri", size=11, bold=True)
        cell_wt_total.alignment = Alignment(horizontal="center")
        cell_wt_total.border = BOTTOM_BORDER
        row += 2  # gap row

    # ===== Section B: Complete Monthly Holdings Timeline =====
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, f"Complete Monthly Holdings Timeline — {period_label}").font = TITLE_FONT
    row += 1

    # Build combined DataFrame from all strategies
    all_holdings = []
    for sname, sdata in strategy_results.items():
        hh = sdata["backtest_result"].holdings_history.copy()
        if hh.empty:
            continue
        hh["strategy"] = sname
        all_holdings.append(hh)

    if all_holdings:
        combined = pd.concat(all_holdings, ignore_index=True)
        combined = combined.sort_values(
            ["date", "strategy", "weight"], ascending=[False, True, False]
        )

        # Headers
        for j, h in enumerate(["Date", "Strategy", "Ticker", "Shares", "Value ($)", "Weight (%)"]):
            cell = ws.cell(row, 1 + j, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for _, rec in combined.iterrows():
            ws.cell(row, 1, rec["date"].strftime("%Y-%m")).font = DATA_FONT
            ws.cell(row, 1).border = THIN_BORDER

            ws.cell(row, 2, rec["strategy"]).font = DATA_FONT
            ws.cell(row, 2).border = THIN_BORDER

            ws.cell(row, 3, rec["ticker"]).font = DATA_FONT
            ws.cell(row, 3).border = THIN_BORDER

            cell_s = ws.cell(row, 4, rec["shares"])
            cell_s.number_format = "#,##0.00"
            cell_s.font = DATA_FONT
            cell_s.alignment = Alignment(horizontal="center")
            cell_s.border = THIN_BORDER

            cell_v = ws.cell(row, 5, rec["value"])
            cell_v.number_format = '"$"#,##0'
            cell_v.font = DATA_FONT
            cell_v.alignment = Alignment(horizontal="center")
            cell_v.border = THIN_BORDER

            cell_w = ws.cell(row, 6, rec["weight"])
            cell_w.number_format = "0.00%"
            cell_w.font = DATA_FONT
            cell_w.alignment = Alignment(horizontal="center")
            cell_w.border = THIN_BORDER

            row += 1

    # ===== Section C: Position Changes =====
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, f"Position Changes — {period_label}").font = TITLE_FONT
    row += 1

    all_changes = []
    for sname, sdata in strategy_results.items():
        hh = sdata["backtest_result"].holdings_history
        if hh.empty:
            continue
        changes = _derive_position_changes(hh, sname)
        if not changes.empty:
            all_changes.append(changes)

    if all_changes:
        combined_changes = pd.concat(all_changes, ignore_index=True)
        combined_changes = combined_changes.sort_values(
            ["date", "strategy", "action", "ticker"], ascending=[False, True, True, True]
        )

        for j, h in enumerate(["Date", "Strategy", "Ticker", "Action", "Weight"]):
            cell = ws.cell(row, 1 + j, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        row += 1

        enter_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        exit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for _, rec in combined_changes.iterrows():
            fill = enter_fill if rec["action"] == "ENTER" else exit_fill

            ws.cell(row, 1, rec["date"].strftime("%Y-%m")).font = DATA_FONT
            ws.cell(row, 1).border = THIN_BORDER

            ws.cell(row, 2, rec["strategy"]).font = DATA_FONT
            ws.cell(row, 2).border = THIN_BORDER

            ws.cell(row, 3, rec["ticker"]).font = DATA_FONT
            ws.cell(row, 3).border = THIN_BORDER

            cell_action = ws.cell(row, 4, rec["action"])
            cell_action.font = Font(name="Calibri", size=11, bold=True)
            cell_action.fill = fill
            cell_action.alignment = Alignment(horizontal="center")
            cell_action.border = THIN_BORDER

            cell_w = ws.cell(row, 5, rec["weight"])
            cell_w.number_format = "0.00%"
            cell_w.font = DATA_FONT
            cell_w.alignment = Alignment(horizontal="center")
            cell_w.border = THIN_BORDER

            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    ws.freeze_panes = "A2"


def write_strategy_definitions_sheet(wb: Workbook):
    """Write a Strategy Definitions sheet at position 1 (after Quick Compare)."""
    ws = wb.create_sheet("Strategy Definitions", 1)

    # Title
    ws.merge_cells("A1:E1")
    ws.cell(1, 1, "Strategy Definitions & Parameters").font = TITLE_FONT

    row = 3
    # Strategy definitions
    strategies_info = [
        (
            "Top-1",
            "Invests 100% in the single largest company by market capitalization. "
            "Rebalances monthly.",
        ),
        (
            "Top-N Equal Weight (N=5)",
            "Invests equally across the top 5 companies by market cap. "
            "Rebalances monthly.",
        ),
        (
            "Log-Momentum(5,6)",
            "Ranks top 5 companies by market cap, scores each by log momentum "
            "log(M(t)/M(t-6)) where M=market cap, 6=lookback months. Allocates "
            "proportionally to positive momentum scores only. 100% concentration "
            "possible when only one company has positive momentum. Falls back to "
            "equal weight when all momentum is negative.",
        ),
    ]

    # Section header
    ws.cell(row, 1, "Strategy").font = HEADER_FONT
    ws.cell(row, 1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row, 2, "Description").font = HEADER_FONT
    ws.cell(row, 2).fill = HEADER_FILL
    for col in range(3, 6):
        ws.cell(row, col).fill = HEADER_FILL
    row += 1

    for sname, desc in strategies_info:
        ws.cell(row, 1, sname).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row, 1).fill = STRATEGY_FILL
        ws.cell(row, 1).border = THIN_BORDER
        ws.cell(row, 1).alignment = Alignment(vertical="top")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        cell_desc = ws.cell(row, 2, desc)
        cell_desc.font = DATA_FONT
        cell_desc.alignment = Alignment(wrap_text=True, vertical="top")
        cell_desc.border = THIN_BORDER
        ws.row_dimensions[row].height = 45
        row += 1

    # Transaction Cost Schedule
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "Transaction Cost Schedule (Round-Trip)").font = Font(
        name="Calibri", size=12, bold=True, color="1F4E79"
    )
    row += 1

    for j, h in enumerate(["Period", "Cost (bps)", "Per-Leg Cost"]):
        cell = ws.cell(row, 1 + j, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    cost_rows = [
        ("Pre-2001", "50 bps", "25 bps (cost_bps / 20000)"),
        ("2001 - 2004", "20 bps", "10 bps"),
        ("2005+", "10 bps", "5 bps"),
    ]
    for period, cost, per_leg in cost_rows:
        ws.cell(row, 1, period).font = DATA_FONT
        ws.cell(row, 1).border = THIN_BORDER
        ws.cell(row, 2, cost).font = DATA_FONT
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).border = THIN_BORDER
        ws.cell(row, 3, per_leg).font = DATA_FONT
        ws.cell(row, 3).alignment = Alignment(horizontal="center")
        ws.cell(row, 3).border = THIN_BORDER
        row += 1

    # Capital Parameters
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "Capital Parameters").font = Font(
        name="Calibri", size=12, bold=True, color="1F4E79"
    )
    row += 1

    for j, h in enumerate(["Parameter", "Value"]):
        cell = ws.cell(row, 1 + j, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    capital_rows = [
        ("Initial Capital", "$10,000"),
        ("Monthly Contributions", "$1,000"),
        ("Rebalance Frequency", "Monthly (month-end)"),
    ]
    for param, val in capital_rows:
        ws.cell(row, 1, param).font = DATA_FONT
        ws.cell(row, 1).border = THIN_BORDER
        ws.cell(row, 2, val).font = DATA_FONT
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def main():
    STRATEGIES = {
        "Top-1": make_top1_fn(),
        "Top-5 EW": make_topn_fn(5),
        # NOTE: Uses fixed default parameters (N=5, k=6). For optimized parameters,
        # see grid_search_results.csv. The dashboard (app.py) reads optimized values.
        "Log-Momentum(5,6)": make_momentum_fn(5, 6),
    }
    # Make STRATEGIES accessible to run_period via closure / global assignment
    globals()["STRATEGIES"] = STRATEGIES

    print("Loading data...")
    data = fetch_all(use_cache=True)
    mcaps = estimate_market_caps(
        data["prices"], data["splits"], data["shares_outstanding"], data["delisted"],
        historical_shares=data.get("historical_shares"),
    )
    rankings = rank_by_market_cap(mcaps)
    rf = data.get("risk_free", pd.Series(dtype=float))
    benchmarks = data["benchmarks"]

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    period_results_cache = {}
    for period_label, start_str in PERIODS.items():
        print(f"Running {period_label}...")
        results = run_period(data, mcaps, rankings, rf, benchmarks, start_str)
        period_results_cache[period_label] = results

        # Sheet name max 31 chars
        sheet_name = period_label.replace("–", "-").replace("(", "").replace(")", "")[:31]
        write_period_sheet(wb, sheet_name, period_label, results)

        # Holdings sheet for this period
        year_tag = start_str[:4]
        holdings_sheet_name = f"Holdings {year_tag}"
        print(f"  Writing {holdings_sheet_name}...")
        write_holdings_sheet(wb, holdings_sheet_name, period_label, results)

    # --- Summary sheet (all periods side by side for Top-5 EW vs S&P 500) ---
    ws = wb.create_sheet("Quick Compare", 0)
    ws.merge_cells("A1:E1")
    ws.cell(1, 1, "Top-5 EW vs S&P 500 Across Periods").font = TITLE_FONT

    ws.cell(3, 1, "Period").font = HEADER_FONT
    ws.cell(3, 1).fill = HEADER_FILL
    for j, h in enumerate(["Top-5 EW CAGR", "S&P 500 CAGR", "Top-5 EW Sharpe", "S&P 500 Sharpe"]):
        cell = ws.cell(3, 2 + j, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, (period_label, start_str) in enumerate(PERIODS.items()):
        print(f"Quick compare: {period_label}...")
        results = period_results_cache[period_label]
        row = 4 + i
        ws.cell(row, 1, period_label).font = Font(name="Calibri", size=11, bold=True)

        t5 = results.get("Top-5 EW", {}).get("metrics", {})
        sp = results.get("S&P 500", {}).get("metrics", {})

        for j, (val, fmt) in enumerate([
            (t5.get("cagr_twr"), "0.00%"),
            (sp.get("cagr_twr"), "0.00%"),
            (t5.get("sharpe_ratio"), "0.0000"),
            (sp.get("sharpe_ratio"), "0.0000"),
        ]):
            cell = ws.cell(row, 2 + j)
            cell.value = val if val is not None else "—"
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    for c in "BCDE":
        ws.column_dimensions[c].width = 18

    # Strategy Definitions sheet (insert at position 1, after Quick Compare)
    print("Writing Strategy Definitions...")
    write_strategy_definitions_sheet(wb)

    # Save
    out_path = Path(__file__).parent.parent / "results" / "strategy_vs_benchmark_comparison.xlsx"
    out_path.parent.mkdir(exist_ok=True)
    wb.save(str(out_path))
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
